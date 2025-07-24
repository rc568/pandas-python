from pathlib import Path
from openpyxl import load_workbook
import csv

hojas_deseadas = ["LISTA", "VENTAS", "LLEGADAS"]

rangos = {
        "LISTA": ("A1", "K"),
        "VENTAS": ("A2", "P"),
        "LLEGADAS": ("A2", "H"),
    }

def gastos_mandril_to_csv(file, output_dir):
    output_dir = Path(output_dir)
    wb = load_workbook(file, data_only=True)

    for hoja in hojas_deseadas:
        ws = wb[hoja]

        last_row = 2

        while ws[f"A{last_row}"].value is not None:
            last_row += 1

        last_row -= 1

        start_cell, end_col = rangos[hoja]
        data = []

        for row in ws[start_cell:f'{end_col}{last_row}']:
            data.append([cell.value for cell in row])
            
        output_dir.mkdir(exist_ok=True, parents=True)

        with open(output_dir / f"{hoja}.csv", "w", newline='', encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(data)
